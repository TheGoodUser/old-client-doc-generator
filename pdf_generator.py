import datetime
import openpyxl # type: ignore
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill # type: ignore
from openpyxl.utils import get_column_letter # type: ignore



def generate_document(*, filename: list, data: list) -> str:
    """
    filename list contains a list with elements:-
        1. ATTENDANCE_REPORT  (filename prefix)
        2. month_name         (filename suffix)
        3. .xlsx              (extension)
    """

    # Create or load the workbook
    file_path = f"reports/{''.join(filename)}"

    # extract the month from the filename list
    month: str = filename[1]
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

    # Clear the worksheet (optional, if you want to overwrite)
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)

    # Define styles
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(bottom=Side(style="thin"))
    head_border = Border(bottom=Side(style="thin"), right=Side(style="thin", color="d9d9d9"))
    header_fill = PatternFill(start_color="cfe2f3", end_color="cfe2f3", fill_type="solid")
    title_fill = PatternFill(start_color="FFC58B", end_color="FFC58B", fill_type="solid")

    overall_details_header_fill = PatternFill(start_color="d9d9d9", end_color="d9d9d9", fill_type="solid")
    overall_details_fill = PatternFill(start_color="fff2cc", end_color="fff2cc", fill_type="solid")

    # Title (Row 1)
    ws.merge_cells("A1:J1")
    ws["A1"] = f"ATTENDANCE REPORT FOR THE MONTH {month.upper()}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = center_alignment
    ws["A1"].fill = title_fill
    ws.row_dimensions[1].height = 40

    # Report Period and Overall Attendance (Row 2)
    ws.merge_cells("A2:B2")
    ws["A2"] = "REPORT PERIOD"
    ws["A2"].font = header_font
    ws["A2"].alignment = center_alignment
    ws["A2"].fill = overall_details_header_fill
    ws.row_dimensions[2].height = 30

    ws.merge_cells("A3:B3")
    ws["A3"] = month.upper()
    ws["A3"].alignment = center_alignment
    ws["A3"].fill = overall_details_fill
    ws.row_dimensions[3].height = 30

    ws.merge_cells("C2:D2")
    # ws["C2"] = "OVERALL ATTENDANCE"
    ws["C2"] = " "
    # ws["C2"].font = header_font
    # ws["C2"].alignment = center_alignment
    ws["C2"].fill = overall_details_header_fill

    ws.merge_cells("C3:D3")
    # ws["C3"] = "1000/1030"
    ws["C3"] = " "
    # ws["C3"].alignment = center_alignment
    ws["C3"].fill = overall_details_fill

    # Empty row (Row 3)
    ws.merge_cells("A4:J4")
    ws["A4"] = " "  # Placeholder to avoid empty merged cell issues
    ws.row_dimensions[4].height = 30

    ws.merge_cells("E2:F2")
    ws["E2"].fill = overall_details_header_fill

    ws.merge_cells("G2:H2")
    ws["G2"].fill = overall_details_header_fill

    ws.merge_cells("I2:J2")
    ws["I2"].fill = overall_details_header_fill

    ws.merge_cells("E3:F3")
    ws["E3"].fill = overall_details_fill

    ws.merge_cells("G3:H3")
    ws["G3"].fill = overall_details_fill

    ws.merge_cells("I3:J3")
    ws["I3"].fill = overall_details_fill


    # Headers (Row 5)
    headers = ["NAME", "EID", "DEPARTMENT", "DATE", "MONTH", "CHECK-IN TIME",
            "CHECK-OUT TIME", "TOTAL WORKING HOURS", "TOTAL WORKING DAYS", "STATUS"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = head_border
        cell.fill = header_fill

    ws.row_dimensions[5].height = 40

    # Add data rows
    for idx, record in enumerate(data, start=6):
        row = [
            record["name"],
            record["eid"],
            record["department"],
            record["date"],
            record["month"],
            record["check_in"],
            record["check_out"],
            record["total_hours"],
            record["total_days"],
            record["status"]
        ]
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=idx, column=col)
            cell.value = value
            cell.alignment = center_alignment
            cell.border = border
        
        ws.row_dimensions[idx].height = 30

    # Footer (after data)
    footer_row = len(data) + 7
    ws.merge_cells(f"A{footer_row}:J{footer_row}")
    ws[f"A{footer_row}"] = "This is a computer generated document."
    ws[f"A{footer_row}"].alignment = center_alignment
    ws[f"A{footer_row}"].font = Font(italic=True)
    ws.row_dimensions[footer_row].height = 30


    # Adjust column widths
    for col in range(1, 12):  # Columns A to K
        column_letter = get_column_letter(col)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    wb.save(file_path)
    print(f"\nExcel file generated: {file_path} at {datetime.datetime.now()}\n")
    return file_path

